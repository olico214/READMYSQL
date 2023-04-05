from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import csv
from openpyxl import Workbook
import openpyxl
from openpyxl import *
import mysql.connector
from mysql.connector import Error
import os
import datetime


class Buscador:
    def __init__(self):
        
        self.root = Tk()
        self.ruta_Global = 0
        self.root.after(0, self.conexion_temporaly)

        self.root.title("ACTUALIZADOR")
        self.root.geometry("560x380")
        
        self.cuerpo= ttk.Frame(self.root)
        self.cuerpo.place(x=0,y=0,height=480, width=480)
        
        
        
        self.historico = ttk.Button(self.root, text="Historico de ventas", command=self.conexion_historico)
        self.historico.place(x=10,y=60,width=200)
        
        self.historico_90 = ttk.Button(self.root, text="INFORME A 90 DIAS", command=self.conexion_90)
        self.historico_90.place(x=280,y=60,width=200)
        
        self.log = tk.Listbox(self.root)
        self.log.place(x=10,y=100,width=500,height=200)

        self.cargar = ttk.Button(self.root, text="Cargar Ruta", command=self.conexion_temporaly)
        self.cargar.place(x=420,y=310)
        
        
        
        
        
        
        
        self.root.mainloop()



    def conexion_temporaly(self):
        
       # Obtener la ruta actual del archivo
        ruta_actual = os.path.dirname(os.path.abspath(__file__))

        # Crear la ruta completa de la carpeta de destino y crearla si no existe
        ruta_destino = os.path.join(ruta_actual, "Ruta_generada")
        


        if not os.path.exists(ruta_destino):
            os.makedirs(ruta_destino)
        

        ruta_destino = os.path.join(ruta_actual, "Ruta_generada")

        
            # Crear un nuevo libro de Excel y obtener la hoja activa
        ruta = messagebox.askyesno("Nueva Ruta","¿ Quiere seleccionar una nueva ruta donde guardar la información ?")

        if ruta == True:
            folder_path = filedialog.askdirectory()
            wb = openpyxl.Workbook()
            hoja_activa = wb.active

            # Insertar la ruta de destino en la celda A1
            hoja_activa['A1'] = folder_path

            # Guardar el libro de Excel
            wb.save(os.path.join(ruta_destino, 'RutaGlobal.xlsx'))
            self.ruta_Global = folder_path
            # Imprimir el contenido de la celda A1
            print("El contenido de la celda A1 es:", self.ruta_Global)
            self.log.insert(tk.END, self.ruta_Global)
        
        if ruta == False:
            try:
            # Abrir el libro de Excel y obtener la hoja activa
                wb = openpyxl.load_workbook(os.path.join(ruta_destino, 'RutaGlobal.xlsx'))
                hoja_activa = wb.active

                # Leer el contenido de la celda A1
                self.ruta_Global = hoja_activa['A1'].value

                # Imprimir el contenido de la celda A1
                print("El contenido de la celda A1 es:", self.ruta_Global)
                self.log.insert(tk.END,  self.ruta_Global)
            except:
                messagebox.showinfo("Info","Debe seleccionar una ruta para continuar")

        
        
        
    def conexion_90(self):
        print("error")
        try:
            conexion = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='',
            db='datos1'
            )
            print("error")
            if conexion.is_connected():
                
                self.log.insert(tk.END, "CONEXION A LA BASE DE DATOS EXITOSA")
                with conexion.cursor() as cursor:
                    fecha_hoy = datetime.date.today()
                    date_today = fecha_hoy.strftime('%Y-%m-%d')

                    fecha_90_dias_atras = fecha_hoy - datetime.timedelta(days=90)
                    date_today_90 = fecha_90_dias_atras.strftime('%Y-%m-%d')
                    
                    
                    print("error")
                    self.log.insert(tk.END, "FECHA DE HOY: " +  date_today)
                    print(date_today_90)
                    self.log.insert(tk.END, "FECHA DE 90 DIAS ATRAS: " +  date_today_90)
                    
                    # En este caso no necesitamos limpiar ningún dato
                    cursor.execute("SELECT Producto, Descrip, sum(Cantidad), cast(FADUA as date)   "    +
                    "from ticketsdetalle where FADUA >= '" + date_today_90 + "' and FADUA <= '" + date_today +"' " +
                    "group by Producto, FADUA " +
                    "order by FADUA ;")
                    
                    

                    # Con fetchall traemos todas las filas
                    data = cursor.fetchall()

                    # Recorrer e imprimir
                    producto = []
                    desc = []
                    cantidad = []
                    fecha = []

                    

                    # Crear el nombre completo del archivo a guardar
                    nombre_archivo = os.path.join(self.ruta_Global, "productos.xlsx")
                    nombre_archivo_2 = os.path.join(self.ruta_Global, "almacen.xlsx")

                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Producto', 'Descripción', 'Venta', 'Fecha'))
                    # Crear una lista de tuplas con los datos
                    datos_filas = [(datos[0], datos[1], datos[2], datos[3]) for datos in data]

                    # Agregar la lista completa como una fila en la hoja de trabajo
                    for fila in datos_filas:
                        hoja.append(fila)
                    try:
                        wb.save(nombre_archivo)
                    except:
                        self.log.insert(tk.END, "CIERRE EL ARCHIVO: PRODUCTOS PARA CONTINUAR")

                with conexion.cursor() as historico:
                    # En este caso no necesitamos limpiar ningún dato
                    historico.execute("select Producto, ExistenciaActual from historicoalmacen")

                    # Con fetchall traemos todas las filas
                    datahist = historico.fetchall()

                    producto_2 = []
                    cntidad_2 = []
                    wb_2 = openpyxl.Workbook()
                    hoja_2 = wb_2.active
                    hoja_2.append(('Producto', 'Cantidad'))
                    # Crear una lista de tuplas con los datos
                    datos_filas_2 = [(datos[0], datos[1]) for datos in datahist]

                    for fila_2 in datos_filas_2:
                        hoja_2.append(fila_2)

                    try:
                        wb_2.save(nombre_archivo_2)
                    except:
                        self.log.insert(tk.END, "CIERRE EL ARCHIVO: ALMACEN PARA CONTINUAR")
                        
                        
                        
                self.log.insert(tk.END, "CONEXION EXITOSA")    
        except Exception as e:
            print(f"Ocurrió un error: {str(e)}")
            self.log.insert(tk.END, "OCURRIO UN ERROR")  
        finally:
            if conexion.is_connected():
                cursor.close()
                conexion.close()
                self.log.insert(tk.END, "PROCESO FINALIZADO")  

            
        
    def conexion_historico(self):
        try:
            conexion = mysql.connector.connect(
                host='localhost',
                port=3306,
                user='root',
                password='',
                db='datos1'
            )

            if conexion.is_connected():
                self.log.insert(tk.END, "CONEXION A LA BASE DE DATOS EXITOSA")
                cursor = conexion.cursor()
                cursor.execute("SELECT Producto, Descrip, sum(Cantidad), cast(FADUA as date) from ticketsdetalle " +
                                "group by Producto, FADUA " +
                                "order by FADUA;")
                
                data = cursor.fetchall()
                
                self.log.insert(tk.END, "DATOS DESCARGADOS CON EXITO")
                
                

                # Crear el nombre completo del archivo a guardar
                nombre_archivo = os.path.join(self.ruta_Global, "historico.xlsx")
                

                wb = openpyxl.Workbook()
                hoja = wb.active
                hoja.append(('Producto', 'Descripción', 'Venta', 'Fecha'))
                # Crear una lista de tuplas con los datos
                datos_filas = [(datos[0], datos[1], datos[2], datos[3]) for datos in data]

                # Agregar la lista completa como una fila en la hoja de trabajo
                for fila in datos_filas:
                    hoja.append(fila)
                try:
                    wb.save(nombre_archivo)
                    self.log.insert(tk.END, "PROCESO FINALIZADO")
                except:
                    self.log.insert(tk.END, "CIERRE EL ARCHIVO: HISTORICO PARA CONTINUAR")
                    
                    
        except Exception as e:
            print(f"Ocurrió un error: {str(e)}")
            self.log.insert(tk.END, "OCURRIO UN ERROR")  
        finally:
            if conexion.is_connected():
                cursor.close()
                conexion.close()
                self.log.insert(tk.END, "PROCESO FINALIZADO") 
                        
                        
            
            
aplicacion1 = Buscador()