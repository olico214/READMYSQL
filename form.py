from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import csv
from openpyxl import Workbook
import openpyxl
from openpyxl import *
import mysql.connector
from mysql.connector import Error
import os
import datetime


class Buscador:
    def __init__(self):
        self.root = Tk()
        self.root.title("ACTUALIZADOR")
        self.root.geometry("480x240")
        
        self.cuerpo= ttk.Frame(self.root)
        self.cuerpo.place(x=0,y=0,height=480, width=480)
        
        
        
        self.historico = ttk.Button(self.root, text="Historico de ventas", command=self.conexion_historico)
        self.historico.place(x=10,y=60,width=200)
        
        self.historico_90 = ttk.Button(self.root, text="INFORME A 90 DIAS", command=self.conexion_90)
        self.historico_90.place(x=230,y=60,width=200)
        
        self.log = tk.Listbox(self.root)
        self.log.place(x=10,y=100,width=450,height=100)
        
        
        
        
        
        
        
        self.root.mainloop()
        
    def conexion_90(self):
        try:
            conexion = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='',
            db='testv1'
            )

            if conexion.is_connected():
                self.log.insert(tk.END, "CONEXION A LA BASE DE DATOS EXITOSA")
                with conexion.cursor() as cursor:
                    fecha_hoy = datetime.date.today()
                    date_today = fecha_hoy.strftime('%Y-%m-%d')

                    fecha_90_dias_atras = fecha_hoy - datetime.timedelta(days=90)
                    date_today_90 = fecha_90_dias_atras.strftime('%Y-%m-%d')
                    
                    
                    print()
                    self.log.insert(tk.END, "FECHA DE HOY: " +  date_today)
                    print(date_today_90)
                    self.log.insert(tk.END, "FECHA DE 90 DIAS ATRAS: " +  date_today_90)
                    
                    # En este caso no necesitamos limpiar ningún dato
                    cursor.execute("SELECT Producto, Descrip, sum(Cantidad), cast(FADUA as date)   "    +
                    "from ticketsdetalle where FADUA >= '" + date_today_90 + "' and FADUA <= '" + date_today +"' " +
                    "group by Producto, FADUA " +
                    "order by FADUA ;")
                    
                    

                    # Con fetchall traemos todas las filas
                    data = cursor.fetchall()

                    # Recorrer e imprimir
                    producto = []
                    desc = []
                    cantidad = []
                    fecha = []

                    ruta_actual = os.path.dirname(os.path.abspath(__file__))

                    # Crear la ruta completa de la carpeta de destino y crearla si no existe
                    ruta_destino = os.path.join(ruta_actual, "ArchivosExcel")
                    if not os.path.exists(ruta_destino):
                        os.makedirs(ruta_destino)

                    # Crear el nombre completo del archivo a guardar
                    nombre_archivo = os.path.join(ruta_destino, "productos.xlsx")
                    nombre_archivo_2 = os.path.join(ruta_destino, "almacen.xlsx")

                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Producto', 'Descripción', 'Venta', 'Fecha'))
                    # Crear una lista de tuplas con los datos
                    datos_filas = [(datos[0], datos[1], datos[2], datos[3]) for datos in data]

                    # Agregar la lista completa como una fila en la hoja de trabajo
                    for fila in datos_filas:
                        hoja.append(fila)
                    try:
                        wb.save(nombre_archivo)
                    except:
                        self.log.insert(tk.END, "CIERRE EL ARCHIVO: PRODUCTOS PARA CONTINUAR")

                with conexion.cursor() as historico:
                    # En este caso no necesitamos limpiar ningún dato
                    historico.execute("select Producto, ExistenciaActual from historicoalmacen")

                    # Con fetchall traemos todas las filas
                    datahist = historico.fetchall()

                    producto_2 = []
                    cntidad_2 = []
                    wb_2 = openpyxl.Workbook()
                    hoja_2 = wb_2.active
                    hoja_2.append(('Producto', 'Cantidad'))
                    # Crear una lista de tuplas con los datos
                    datos_filas_2 = [(datos[0], datos[1]) for datos in datahist]

                    for fila_2 in datos_filas_2:
                        hoja_2.append(fila_2)

                    try:
                        wb_2.save(nombre_archivo_2)
                    except:
                        self.log.insert(tk.END, "CIERRE EL ARCHIVO: ALMACEN PARA CONTINUAR")
                        
                        
                        
                self.log.insert(tk.END, "CONEXION EXITOSA")    
        except Exception as e:
            print(f"Ocurrió un error: {str(e)}")
            self.log.insert(tk.END, "OCURRIO UN ERROR")  
        finally:
                if conexion.is_connected():
                    cursor.close()
                    conexion.close()
                    self.log.insert(tk.END, "PROCESO FINALIZADO")  

            
        
    def conexion_historico(self):
        try:
            conexion = mysql.connector.connect(
                host='localhost',
                port=3306,
                user='root',
                password='',
                db='testv1'
            )

            if conexion.is_connected():
                self.log.insert(tk.END, "CONEXION A LA BASE DE DATOS EXITOSA")
                cursor = conexion.cursor()
                cursor.execute("SELECT Producto, Descrip, sum(Cantidad), cast(FADUA as date) from ticketsdetalle " +
                                "group by Producto, FADUA " +
                                "order by FADUA;")
                
                data = cursor.fetchall()
                
                self.log.insert(tk.END, "DATOS DESCARGADOS CON EXITO")
                
                ruta_actual = os.path.dirname(os.path.abspath(__file__))

                # Crear la ruta completa de la carpeta de destino y crearla si no existe
                ruta_destino = os.path.join(ruta_actual, "ArchivosExcel")
                if not os.path.exists(ruta_destino):
                    os.makedirs(ruta_destino)

                # Crear el nombre completo del archivo a guardar
                nombre_archivo = os.path.join(ruta_destino, "historico.xlsx")
                

                wb = openpyxl.Workbook()
                hoja = wb.active
                hoja.append(('Producto', 'Descripción', 'Venta', 'Fecha'))
                # Crear una lista de tuplas con los datos
                datos_filas = [(datos[0], datos[1], datos[2], datos[3]) for datos in data]

                # Agregar la lista completa como una fila en la hoja de trabajo
                for fila in datos_filas:
                    hoja.append(fila)
                try:
                    wb.save(nombre_archivo)
                    self.log.insert(tk.END, "PROCESO FINALIZADO")
                except:
                    self.log.insert(tk.END, "CIERRE EL ARCHIVO: HISTORICO PARA CONTINUAR")
                    
                    
        except Exception as e:
            print(f"Ocurrió un error: {str(e)}")
            self.log.insert(tk.END, "OCURRIO UN ERROR")  
        finally:
                if conexion.is_connected():
                    cursor.close()
                    conexion.close()
                    self.log.insert(tk.END, "PROCESO FINALIZADO") 
                        
                        
            
            
aplicacion1 = Buscador()